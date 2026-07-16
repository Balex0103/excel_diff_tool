import pandas as pd

def load_excel(file_path):
    """Beolvassa az Excel fájlt, és automatikusan kitakarítja a felesleges üres oszlopokat."""
    df = pd.read_excel(file_path)
    
    # Csak azokat az oszlopokat tartjuk meg, amikben van tényleges adat (nem teljesen üresek)
    df = df.dropna(how='all', axis=1)
    
    # Eltávolítjuk a névtelen "Unnamed" oszlopokat is
    df = df.loc[:, ~df.columns.str.contains('^Unnamed:', na=False)]
    return df


def compare_dataframes(df_old: pd.DataFrame, df_new: pd.DataFrame, key_cols: list[str]):
    """
    Szupergyors, vektorizált összehasonlítás a megadott kulcsoszlopok alapján.
    Visszaadja a csak régiben lévő, csak újban lévő, és a módosult értékek listáját.
    """
    # Másolatokat készítünk, hogy ne sérüljenek az eredeti adatok
    df_o = df_old.copy()
    df_n = df_new.copy()

    # Kulcsoszlopok beállítása indexként a gyors párosításhoz
    df_o.set_index(key_cols, inplace=True, drop=False)
    df_n.set_index(key_cols, inplace=True, drop=False)

    # 1. CSAK RÉGIBEN LÉVŐK (Törölt sorok)
    only_old = df_o[~df_o.index.isin(df_n.index)].copy()

    # 2. CSAK ÚJBAN LÉVŐK (Új sorok)
    only_new = df_n[~df_n.index.isin(df_o.index)].copy()

    # 3. MÓDOSULT ÉRTÉKEK KERESÉSE
    # Csak azokat a sorokat vizsgáljuk, amik mindkét fájlban megtalálhatóak
    common_indices = df_o.index.intersection(df_n.index)
    df_o_common = df_o.loc[common_indices]
    df_n_common = df_n.loc[common_indices]

    # Összehasonlítandó oszlopok meghatározása (kulcsoszlopok nélkül)
    compare_cols = [col for col in df_o_common.columns if col not in key_cols and col in df_n_common.columns]

    changed_records = []

    # Oszloponként futtatjuk a vektorizált összehasonlítást a maximális sebességért
    for col in compare_cols:
        # Kikeressük az eltéréseket, ügyelve arra, hogy a NaN == NaN egyenlőség teljesüljön
        diff_mask = (df_o_common[col] != df_n_common[col]) & ~(df_o_common[col].isna() & df_n_common[col].isna())
        
        if diff_mask.any():
            diff_rows_o = df_o_common[diff_mask]
            diff_rows_n = df_n_common[diff_mask]
            
            for idx in diff_rows_o.index:
                record = {}
                # Kulcs értékek visszanyerése (kezelve az összetett kulcsokat is)
                if len(key_cols) == 1:
                    record[key_cols[0]] = idx
                else:
                    for k_idx, k_col in enumerate(key_cols):
                        record[k_col] = idx[k_idx]
                
                record["oszlop"] = col
                record["regi_ertek"] = diff_rows_o.loc[idx, col]
                record["uj_ertek"] = diff_rows_n.loc[idx, col]
                changed_records.append(record)

    if changed_records:
        changed_df = pd.DataFrame(changed_records)
    else:
        # Ha nincs változás, üres DataFrame-et adunk vissza a megfelelő oszlopfejlécekkel
        changed_df = pd.DataFrame(columns=key_cols + ["oszlop", "regi_ertek", "uj_ertek"])

    # Visszaállítjuk az indexeket a tiszta Excel riport kimenet érdekében
    only_old.reset_index(drop=True, inplace=True)
    only_new.reset_index(drop=True, inplace=True)

    return only_old, only_new, changed_df


def write_report(only_old, only_new, changed_df, output_path):
    """Létrehozza a riportot professzionális színezéssel és automatikus oszlopszélességgel."""
    # 1. Alap adatok kiírása Excel-be külön fülekre
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        changed_df.to_excel(writer, sheet_name="Modosult_ertekek", index=False)
        only_old.to_excel(writer, sheet_name="Csak_regiben", index=False)
        only_new.to_excel(writer, sheet_name="Csak_ujban", index=False)

    # 2. Formázás openpyxl segítségével
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    wb = load_workbook(output_path)

    # Professzionális, lágy pasztell színek (sötét betűkkel a tökéletes olvashatóságért)
    red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    red_font = Font(color="C00000", name="Calibri", size=11)

    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    green_font = Font(color="375623", name="Calibri", size=11)

    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    yellow_font = Font(color="7F6000", name="Calibri", size=11)

    # Fejléc stílus (Sötétszürke háttér, fehér félkövér betűk)
    header_fill = PatternFill(start_color="595959", end_color="595959", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)

    def format_sheet(sheet_name, fill, font, col_count):
        if sheet_name not in wb.sheetnames:
            return
        ws = wb[sheet_name]
        
        # Ha a munkalap üres (csak a fejléc van benne)
        if ws.max_row <= 1:
            return

        # Fejléc formázása
        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font

        # Adatsorok formázása (Csak a tényleges adatoszlopokig színez!)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=col_count):
            for cell in row:
                cell.fill = fill
                cell.font = font

        # Oszlopszélességek automatikus igazítása (Auto-fit)
        for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=col_count):
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value) if cell.value is not None else ""
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Formázások végrehajtása az egyes lapokon a megfelelő színekkel
    format_sheet("Modosult_ertekek", yellow_fill, yellow_font, len(changed_df.columns))
    format_sheet("Csak_regiben", red_fill, red_font, len(only_old.columns))
    format_sheet("Csak_ujban", green_fill, green_font, len(only_new.columns))

    wb.save(output_path)